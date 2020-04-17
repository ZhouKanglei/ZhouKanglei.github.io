import os
import random
import openpyxl
import time
import requests
from lxml import etree
from bs4 import BeautifulSoup
import re


def CNKI(filename, title_input, keyword_input, theme_input, author_input, unit_input, content_input, publishTimeBegine_input, publishTimeEnd_input):
    url = 'http://yuanjian.cnki.net/Search/ListResult'
    user_agent = {
      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
    }

    html = ''
    results = []
    break_flag = 0

    time_start = time.time()
    title_link_result = openpyxl.Workbook()
    sheet = title_link_result.active
    sheet.title = 'result'

    col_name = ['title', 'keywords', 'authors', 'journal', 'journal_type', 'year', 'issue', 'download_num', 'cited_num', 'link']
    sheet.append(col_name)

    for i in range(1, 100):
        for j in range(i, i + 1):

            try:
                paramas = {
                  'searchType': 'MulityTermsSearch',
                  'ArticleType': '1',
                  'ParamIsNullOrEmpty': 'true',
                  'Islegal': 'false',
                  'Content': content_input,
                  'Title': title_input,
                  'Author': author_input,
                  'Unit': unit_input,
                  'Keyword': keyword_input,
                  'Theme': theme_input,
                  'publishtimebegin': publishTimeBegine_input,
                  'publishtimeend': publishTimeEnd_input,
                  'Type': '1',
                  'Order': '2',
                  'Page': str(j)
                }

                res = requests.get(url, params = paramas)
                # print(res.status_code)
                soup = BeautifulSoup(res.text, 'html.parser')
                items = soup.find_all('div', class_='list-item')


                for item in items:
                    title = item.find('a')['title']
                    try:
                        try:
                            keywords = item.find('div', class_='info').find('p',class_='info_left left').find_all('a')[0]['data-key']
                            article_info = item.find('p', class_='source').text.replace('\n',' ')
                            
                            article_info = article_info.replace('...','等')
                            article_info_array = article_info.split(' ')
                            
                            for num_empty in range(article_info_array.count('')):
                                article_info_array.remove('')

                            
                            article_info_array_len = len(article_info_array)
                            article_type = article_info_array[article_info_array_len - 1]
                            article_year_issue = article_info_array[article_info_array_len - 2]
                            article_year_issue = article_year_issue.replace('年',' ')
                            article_year_issue = article_year_issue.replace('期','')
                            article_year_issue = article_year_issue.split(' ')
                            
                            if len(article_year_issue) == 2:
                                article_year = article_year_issue[0]
                                article_issue = article_year_issue[1]
                            else:
                                article_year = article_year_issue[0]
                                article_issue = ''
                            article_journal = article_info_array[article_info_array_len - 3]
                            
                            article_authors = ''
                            for article_author_num in range(article_info_array_len - 3):
                                if article_author_num == 0:
                                    article_authors = article_info_array[article_author_num]
                                else:
                                    article_authors = article_authors + ', ' + article_info_array[article_author_num]
                            
                            download_num = re.findall('\d{1,10}',item.find('div',class_='info').find('span', class_='time1').text)[0]
                            cited_num = re.findall('\d{1,10}',item.find('div',class_='info').find('span', class_='time2').text)[0]
                            CNKI_link = item.find('a')['href']

                            link_page = requests.get(CNKI_link)
                            link_page_code = etree.HTML(link_page.text)

                            # 作者
                            article_authorss = []
                            article_authorss = link_page_code.xpath("//*[@id='content']/div[2]/div[3]/a/text()")
                            
                            authors = ''
                            if len(article_authorss) != 0:
                                for article_authorss_num in range(len(article_authorss)):
                                    if article_authorss_num == 0:
                                        authors = authors + article_authorss[article_authorss_num]
                                    else:
                                        authors = authors + ', ' + article_authorss[article_authorss_num]
                                article_authors = authors    
                            

                        except IndexError: 
                            keywords = 'None'
                            article_info = 'None'
                            article_journal = 'None'
                            article_authors = 'None'
                            article_year = 'None'
                            article_issue = 'None'
                            article_type = 'None' 
                            download_num = 'None'
                            cited_num = 'None'
                            CNKI_link = 'None'
                    except AttributeError:
                        keywords = 'None'
                        article_info = 'None'
                        article_journal = 'None'
                        article_authors = 'None'
                        article_year = 'None'
                        article_issue = 'None'
                        article_type = 'None'
                        download_num = 'None'
                        cited_num = 'None'
                        CNKI_link = 'None'

                    if html.find(title) == -1:
                        html = html + res.text     

                    log = [title, keywords, article_authors, article_journal, article_type, article_year, article_issue, download_num, cited_num, CNKI_link]
                    if results == '':
                        results.append(log)
                        sheet.append(log)
                        print(log)
                    else:
                        if log not in results:
                            results.append(log)
                            sheet.append(log)
                            print(log)
                        else:
                            break_flag = 1

                time_end = time.time()

                if len(results) == 0:
                    break_flag = 1
                    break
                print('成功爬取：', len(results), '条  耗时：', str(time_end-time_start), '秒')


                time.sleep(0)
            except TimeoutError:
                print('Time out!!!')
                title_link_result.save(filename + '_break.xlsx')
                break
            time.sleep(0)

        if break_flag == 1:
            break
            
    title_link_result.save(filename + '_all.xlsx')
    file = open(filename + '.html', 'w', encoding = 'utf-8')
    file.write(html)
    file.close()

    return html

def readTable(tableName):
    wb = openpyxl.load_workbook(tableName)
    ws = wb['Sheet1']
    ws_rows_len = ws.max_row         #行数
    ws_columns_len = ws.max_column    #列数

    info = []
    for stu_num in range(ws_rows_len):
        if stu_num != 0:
            info_item = [ws.cell(stu_num + 1, 2).value, '河南师范大学计算机与信息工程学院']
            info.append(info_item)
    
    return info


if __name__ == '__main__':
    # tableName = 'E:\\Office_Files\\Excel\\计科二班联系方式.xlsx'
    # info = readTable(tableName)

    # filename = 'result'
    # title = ''
    # keyword = ''
    # theme = ''
    # content = ''

    # for info_item in  info:
    #     author = info_item[0]
    #     unit = info_item[1]
    #     publishTimeBegine = '2016年'
    #     publishTimeEnd = '2020年'
    #     print(author, ' ', unit)
    #     results = CNKI(filename, title, keyword, theme, author, unit, content, publishTimeBegine, publishTimeEnd)


    filename = 'result'
    title = ''
    keyword = ''
    theme = ''
    content = ''
    unit = ''
    author = '周康垒'
    publishTimeBegine = ''
    publishTimeEnd = ''
    results = CNKI(filename, title, keyword, theme, author, unit, content, publishTimeBegine, publishTimeEnd)
    